import io


from asyncio import get_event_loop
from botocore.exceptions import ClientError
from openpyxl.reader.excel import load_workbook

from celery_config import app
from src.library.messages import responses
from src.helpers.update_book_read_only import update_book_read_only
from src.helpers.update_book_filename import update_book_filename
from src.library.services.s3_service import S3Service


loop = get_event_loop()


@app.task()
def upload_book_file_to_s3(
        file: bytes,
        filename: str,
        book_id: int,

) -> str:
    s3_service = S3Service()

    try:
        s3_service.upload_file(file=file, filename=filename)
        loop.run_until_complete(
            update_book_filename(
                book_id=book_id,
                filename=filename,
            )
        )

        return responses.S3_BOOK_FILE_SUCCESSFULLY_UPLOADED
    except ClientError as error:
        return str(error)


@app.task()
def parse_file_and_add_books_to_denied_books(denied_books_bytes: bytes):
    excel_bytes_io = io.BytesIO(denied_books_bytes)
    workbook = load_workbook(excel_bytes_io)

    books_title = []
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        books_title.extend(
            [
                cell.value
                for row in sheet.iter_rows(
                    min_row=1,
                    max_row=sheet.max_row,
                    min_col=1,
                    max_col=sheet.max_column
                )
                for cell in row
            ]
        )

    loop.run_until_complete(
        update_book_read_only(
            books_title=books_title
        )
    )

    return responses.XLSX_BOOK_DENIED_FILE_SUCCESSFULLY_PARSED
